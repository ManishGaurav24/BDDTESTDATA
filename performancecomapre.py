import json
import time
import boto3
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from io import BytesIO
import os
from openpyxl.styles import PatternFill, Font, Border, Side

# Initialize S3 client
AWS_ACCESS_KEY_ID = os.getenv("aws_access_key_id")
AWS_SECRET_ACCESS_KEY = os.getenv("aws_secret_access_key")
s3_client = boto3.client(
    "s3",
    aws_access_key_id=AWS_ACCESS_KEY_ID,
    aws_secret_access_key=AWS_SECRET_ACCESS_KEY
    # aws_session_token=AWS_SESSION_TOKEN,
)  # Replace with your actual S3 bucket name

# Define a border style
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def apply_border_to_table(sheet, min_row, max_row, min_col, max_col):
    """Apply borders to all cells in the specified range."""
    for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border


def metric_to_sheet_name(metric):
    """Convert metric names to descriptive sheet names."""
    metric_to_name_map = {
        "percentiles1": "50th Percentile",
        "percentiles2": "75th Percentile",
        "percentiles3": "95th Percentile",
        "percentiles4": "99th Percentile"
    }
    return metric_to_name_map.get(metric, metric)


def compare_stats(stats1, stats2, metric, sheet, row, summary):
    """Compare a specific metric between two JSON objects and write the result to the Excel sheet."""
    value1_ok = stats1[metric]["ok"]
    value2_ok = stats2[metric]["ok"]
    value1_ko = stats1[metric]["ko"]
    value2_ko = stats2[metric]["ko"]

    diff_ok = value2_ok - value1_ok
    diff_ko = value2_ko - value1_ko

    safe_value1_ok = value1_ok if value1_ok != 0 else 1e-10
    diff_delta = (value2_ok - safe_value1_ok) / safe_value1_ok * 100
    rounded_diff_delta = round(diff_delta, 2)
    formatted_value = f"{rounded_diff_delta:.2f}%"

    # Keep track of which run is better based on the delta
    if rounded_diff_delta < 0:
        summary['faster'] += 1  # File 2 is faster
    elif rounded_diff_delta > 0:
        summary['slower'] += 1  # File 1 is slower

    # Write the values to the Excel sheet
    sheet.cell(row=row, column=1, value=stats1['name'])
    sheet.cell(row=row, column=2, value=value1_ok)
    sheet.cell(row=row, column=3, value=value2_ok)
    sheet.cell(row=row, column=4, value=diff_ok)

    # Color the cell based on the delta
    cell = sheet.cell(row=row, column=5, value=formatted_value)
    if rounded_diff_delta < 0:
        cell.fill = PatternFill(start_color="0c880e", end_color="0c880e", fill_type="solid")  # Green for Fast
    elif rounded_diff_delta > 0:
        cell.fill = PatternFill(start_color="f4eb36", end_color="f4eb36", fill_type="solid")  # Yellow for Slow

    sheet.cell(row=row, column=6, value=value1_ko)
    sheet.cell(row=row, column=7, value=value2_ko)
    sheet.cell(row=row, column=8, value=diff_ko)

    # Apply border to the current row
    apply_border_to_table(sheet, min_row=row, max_row=row, min_col=1, max_col=8)


def create_sheet(wb, sheet_name, metric):
    """Create a new sheet for a specific metric and set up headers."""
    sheet = wb.create_sheet(title=sheet_name)

    # Define header colors
    header_fill = PatternFill(start_color="51ABD2", end_color="51ABD2", fill_type="solid")
    bold_font = Font(bold=True)

    # Define a different color for the "Delta" columns
    special_fill = PatternFill(start_color="d3d0d0", end_color="d3d0d0", fill_type="solid")

    # Write headers
    headers = ["Transaction Name", f"Execution1_OK Value {metric}", f"Execution2_OK Value {metric}",
               f"Delta(Execution2 value - Execution1 value)", f"Delta Percentage %",
               f"Execution1_KO Value {metric}", f"Execution2_KO Value {metric}",
               f"Delta(Execution2 value - Execution1 value)"]
    for col, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        if "Delta" in header:
            cell.fill = special_fill
        else:
            cell.fill = header_fill
        cell.font = bold_font

    # Apply border to header row
    apply_border_to_table(sheet, min_row=1, max_row=1, min_col=1, max_col=8)

    return sheet


def add_summary(sheet, row, summary):
    """Add a summary at the end of the sheet indicating which run is better."""
    transactions_faster = summary.get('faster', 0)
    transactions_slower = summary.get('slower', 0)
    if transactions_slower > transactions_faster:
        summary_text = "Execution 1 performed better overall."
    elif transactions_faster > transactions_slower:
        summary_text = "Execution 2 performed better overall."
    else:
        summary_text = "Both runs performed equally well."

    note = "These are system generated observations. Please review Manually for a complete analysis"
    bold_font = Font(bold=True)

    blue_fill = PatternFill(start_color="51ABD2", end_color="51ABD2", fill_type="solid")
    green_fill = PatternFill(start_color="65E823", end_color="65E823", fill_type="solid")

    sheet.append(["No.Of.Transactions Executed Faster than last run:", transactions_faster])
    sheet.append(["No.Of.Transactions Executed Slower than last run:", transactions_slower])
    sheet.append(["Overall Observation:", summary_text])
    sheet.append(["*Note:", note])

    # Apply blue fill to static text and green fill to result values
    for i, row in enumerate(sheet.iter_rows(min_row=sheet.max_row - 3, max_row=sheet.max_row, min_col=1, max_col=2)):
        for j, cell in enumerate(row):
            cell.font = bold_font
            if j == 0:
                cell.fill = blue_fill
            else:
                cell.fill = green_fill

    # Apply border to the summary section
    apply_border_to_table(sheet, min_row=sheet.max_row - 3, max_row=sheet.max_row, min_col=1, max_col=2)


def compare_json(json1, json2, metrics):
    # Create an Excel workbook
    wb = Workbook()

    for metric in metrics:
        sheet_name = metric_to_sheet_name(metric)
        sheet = create_sheet(wb, sheet_name, metric)

        summary = {'faster': 0, 'slower': 0}

        def compare_section(json1_section, json2_section, start_row):
            compare_stats(json1_section["stats"], json2_section["stats"], metric, sheet, start_row, summary)
            return start_row + 1

        next_row = 2
        next_row = compare_section(json1, json2, next_row)

        for key in json1.get('contents', {}):
            if key in json2.get('contents', {}):
                next_row = compare_section(json1['contents'][key], json2['contents'][key], next_row)
            else:
                sheet.cell(row=next_row, column=1, value=f"Key '{key}' not found in the second Execution.")
                next_row += 1

        add_summary(sheet, next_row, summary)

    del wb["Sheet"]  # Remove default sheet

    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    performance_comparator_bucket = os.getenv("aws_performance_comparator_bucket")
    ts = str(int(round(time.time())))
    s3_key = f"output_{ts}.xlsx"
    response = s3_client.put_object(
        Bucket=performance_comparator_bucket,
        Key=s3_key,
        Body=excel_buffer.getvalue(),
        ContentType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

    # Check the response and return the download URL
    status = response.get("ResponseMetadata", {}).get("HTTPStatusCode")
    if status == 200:
        url = f"https://{performance_comparator_bucket}.s3.amazonaws.com/{s3_key}"
        return url
    else:
        raise Exception("Failed to upload file to S3")
